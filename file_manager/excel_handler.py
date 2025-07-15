from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework.response import Response
from rest_framework import status
from api.responses import TG_EXCEL_FILE_INVALID, TG_EXCEL_FILE_REQUIRED_COL, TG_EXCEL_FILE_LIMIT_1000


class ExcelHandler:
    @staticmethod
    def generate_excel(headers, rows, **kwargs):
        wb = Workbook()
        ws = wb.active
        if not kwargs.get('file_name'):
            ws.title = "Excel.xlsx"
        else:
            ws.title = kwargs['file_name']
        thin_border = Border(
            left=Side(style='thin', color='777777'),
            right=Side(style='thin', color='777777'),
            top=Side(style='thin', color='777777'),
            bottom=Side(style='thin', color='777777')
        )
        header_font = Font(size=10)
        header_fill = PatternFill(start_color='ffcc00', end_color='ffcc00', fill_type='solid')
        check_col_number = 0
        check_field = kwargs.get('check_field')
        if check_field:
            if check_field in headers:
                check_col_number = headers.index(check_field) + 1
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for row in rows:
            ws.append(row)
        ws.row_dimensions[1].height = 25  # MEH: Set row height for header row
        for row in ws.iter_rows(min_row=1):  # MEH: Example: highlight without order users
            for cell in row:  # MEH: Alignment Center all cel
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            if check_col_number: # MEH: Chack around chosen field and mark those row with fill color
                if row[check_col_number-1].value == '-' or not row[check_col_number-1].value:
                    for cell in row:
                        cell.fill = PatternFill(start_color='dedede', end_color='dedede', fill_type='solid')
        for col_num, column_cells in enumerate(ws.columns, 1):  # Auto col Width
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(col_num)].width = length + 2
        res = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')  # MEH: Prepare HTTP response
        res['Content-Disposition'] = 'attachment; filename=users.xlsx'
        wb.save(res)
        return res

    @staticmethod
    def import_excel(file, allowed_fields, required_field=None, **kwargs):
        try: # MEH: Check for invalid file
            wb = load_workbook(filename=file)
            sheet = wb.active
        except Exception as e:
            return Response({'detail': TG_EXCEL_FILE_INVALID + str(e)}, status=status.HTTP_400_BAD_REQUEST)
        if sheet.max_row - 1 > 1000:
            return Response({'detail': TG_EXCEL_FILE_LIMIT_1000}, status=status.HTTP_400_BAD_REQUEST)
        header = [cell.value for cell in sheet[1]] # MEH: Header first row
        if required_field and not all(col in header for col in required_field):
            return Response({'detail': TG_EXCEL_FILE_REQUIRED_COL + str(required_field)}, status=status.HTTP_400_BAD_REQUEST)
        nested_fields = kwargs.get('nested_fields', [])
        nested_field_category = kwargs.get('nested_field_category')
        extra_fields = kwargs.get('extra_fields', {})
        data_list = []
        for row in list(sheet.iter_rows(min_row=2, values_only=True))[:1000]:
            if not any(row):
                continue  # MEH: Skip empty rows
            row_data = dict(zip(header, row))
            nested_data = {k: row_data.pop(k) for k in list(row_data) if k in nested_fields} # MEH: Pop Nested field from each row
            cleaned_data = {k: v for k, v in row_data.items() if k in allowed_fields} # MEH: Get other clean data from each row
            if nested_field_category:
                cleaned_data.pop(nested_field_category, None) # MEH: Drop data if cell header is nested category! (It's Blocked in this way)
            if nested_data and nested_field_category:
                cleaned_data[nested_field_category] = nested_data
            cleaned_data.update(extra_fields) # MEH: Add extra fields if provided
            data_list.append(cleaned_data) # MEH: Clean Data to check validation in serializer later
        return data_list
