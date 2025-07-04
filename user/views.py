from rest_framework import status, filters
from rest_framework.permissions import IsAuthenticated, IsAdminUser, AllowAny
from rest_framework.response import Response
from api.permissions import ApiAccess, IsNotAuthenticated
from rest_framework.decorators import action
from django_filters.rest_framework import DjangoFilterBackend
from .models import User, Role, Introduction, Address
from django.db.models import Subquery, OuterRef
from .serializers import (UserSignUpSerializer, UserSignInSerializer, UserSerializer,
                          UserImportGetDataSerializer, UserImportSetDataSerializer,
                          UserProfileSerializer,UserRoleSerializer,
                          UserKeySerializer, UserAccountingSerializer, AddressSerializer, IntroductionSerializer, RoleSerializer)
from rest_framework_simplejwt.views import TokenRefreshView, TokenVerifyView
from .filters import CustomerQueryFilter, CustomerFilter
from django.core.exceptions import ObjectDoesNotExist
from rest_framework.exceptions import NotFound
from drf_spectacular.utils import extend_schema, OpenApiResponse
from api.responses import *
from api.mixins import CustomMixinModelViewSet
from api.serializers import BulkDeleteSerializer
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


# MEH: Get Refresh Token
@extend_schema(tags=["Auth"])
class CustomTokenRefreshView(TokenRefreshView):
    pass


# MEH: Verify Access Token
@extend_schema(tags=["Auth"])
class CustomTokenVerifyView(TokenVerifyView):
    pass


# MEH: User List view set
@extend_schema(tags=['Users'])
class UserViewSet(CustomMixinModelViewSet):
    queryset = User.objects.prefetch_related('user_profile')
    serializer_class = UserSerializer
    # MEH: Handle Access for Employee (List, Obj, and per default and custom @action)
    permission_classes = [ApiAccess]
    filterset_class = CustomerFilter
    filter_backends = [
        DjangoFilterBackend,
        CustomerQueryFilter,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    # MEH: Get search query
    search_fields = ['first_name', 'last_name', 'phone_number']
    ordering_fields = ['date_joined', 'order_count', 'last_order_date']
    required_api_keys = {
        'list': 'get_users',
        'retrieve': 'get_users',
        'get_by_phone': 'get_users',
        'create': 'create_user',
        'update': 'update_user',
        'partial_update': 'update_user',
        'profile': 'update_user',
        'destroy': 'delete_user',
        'activation': 'active_user',
        'download_user_list': 'download_user_list',
        'import_users': 'import_user_list',
        'address_list': 'get_addresses',
        'address_detail': 'get_addresses',
        'address_create': 'create_address',
    }

    # MEH: Get province from default Address if there is any & Use it on filter User Province
    def get_queryset(self):
        default_province_id = Subquery(
            Address.objects.filter(user=OuterRef('pk'), is_default=True).values('province')[:1]
        )
        return User.objects.annotate(default_province_id=default_province_id)

    # MEH: Override get single user (with ID or phone_number) | Access check after In has_object_permission
    def get_object(self, *args, **kwargs):
        queryset = self.get_queryset()
        if not self.kwargs.get('phone_number'):
            lookup_value = self.kwargs.get(self.lookup_field, '')
            try:
                obj = queryset.get(pk=int(lookup_value))
                return obj
            except ObjectDoesNotExist:
                raise NotFound(TG_USER_NOT_FOUND_BY_ID)
            except ValueError:
                raise NotFound(TG_EXPECTED_ID_NUMBER)
        lookup_value = self.kwargs.get('phone_number')
        try:
            return queryset.get(phone_number=lookup_value)
        except ObjectDoesNotExist:
            raise NotFound(TG_USER_NOT_FOUND_BY_PHONE)

    # MEH: User Sign Up action (POST) for customer
    @extend_schema(tags=['Auth'])
    @action(detail=False, methods=['post'],
            url_path='sign-up', serializer_class=UserSignUpSerializer,
            permission_classes=[IsNotAuthenticated])
    def sign_up(self, request):
        return self.custom_create(request.data)

    # MEH: User Sign In action (POST) and get Access & Refresh Token
    @extend_schema(tags=['Auth'])
    @action(detail=False, methods=['post'],
            url_path='sign-in', serializer_class=UserSignInSerializer,
            permission_classes=[IsNotAuthenticated])
    def sign_in(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        return Response(serializer.validated_data, status=status.HTTP_200_OK)

    # MEH: Get User by phone_number
    @action(detail=False, methods=['get'],
            url_path='by-phone/(?P<phone_number>\\d{11})')
    def get_by_phone(self, request, phone_number=None):
        queryset = self.get_object(phone_number=phone_number)
        return self.custom_get(queryset)

    # MEH: Bulk Create for user
    @extend_schema(
        request={
            'multipart/form-data': {
                'type': 'object',
                'properties': {
                    'excel_file': {'type': 'string', 'format': 'binary'},
                    'role': {'type': 'integer'}
                },
                'required': ['excel_file', 'role'],
            }
        },
    )
    @action(detail=False, methods=['get', 'post'],
            url_path='import_users', serializer_class=UserImportGetDataSerializer, filter_backends=[], pagination_class=None)
    def import_users(self, request):
        if request.method == 'GET':
            return Response({'detail': list(UserImportSetDataSerializer().get_fields().keys())}, status=status.HTTP_200_OK)
        check_serializer = self.get_serializer(data=request.data)
        check_serializer.is_valid(raise_exception=True)
        # todo: reassemble Excel file later
        excel_file = check_serializer.validated_data['excel_file']
        role = check_serializer.validated_data['role']
        try:
            wb = load_workbook(filename=excel_file)
            sheet = wb.active
        except Exception as e:
            return Response({'detail': TG_EXCEL_FILE_INVALID + str(e)}, status=status.HTTP_400_BAD_REQUEST)
        # Read header row
        header = [cell.value for cell in sheet[1]]
        required_columns = ['phone_number', 'first_name']
        if not all(col in header for col in required_columns):
            return Response({'detail': TG_EXCEL_FILE_REQUIRED_COL + str(required_columns)}, status=status.HTTP_400_BAD_REQUEST)
        user_data_list = []
        profile_fields = [
            name for name, field in UserProfileSerializer().get_fields().items()
            if not getattr(field, 'read_only', False)
        ]
        allowed_fields = set(UserSerializer().get_fields().keys())
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(header, row))
            if not any(row_data.values()):
                continue  # MEH: Skip empty rows
            if role:
                row_data['role'] = role.pk
            user_profile_data = {k: row_data.pop(k) for k in list(row_data) if k in profile_fields}
            cleaned_user_data = {k: v for k, v in row_data.items() if k in allowed_fields}
            cleaned_user_data.pop('user_profile', None)
            if user_profile_data:
                cleaned_user_data['user_profile'] = user_profile_data
            user_data_list.append(cleaned_user_data)
        self.serializer_class = UserImportSetDataSerializer
        res = self.custom_create(user_data_list, many=True)
        self.serializer_class = UserImportGetDataSerializer
        return res


    # MEH: Get User Profile information and Update it with ID
    @action(detail=True, methods=['get', 'put', 'patch'],
            url_path='profile', serializer_class=UserProfileSerializer)
    def profile(self, request, pk=None):
        user = self.get_object(pk=pk)
        queryset = user.user_profile
        if request.method in ['PUT', 'PATCH']:
            return self.custom_update(queryset, request.data, partial=(request.method == 'PATCH'))
        return self.custom_get(queryset)

    # MEH: Get User Key information (Update not work)
    @action(detail=True, methods=['get'],
            url_path='key', serializer_class=UserKeySerializer)
    def key(self, request, pk=None):
        queryset = self.get_object(pk=pk)
        return self.custom_get(queryset)

    # MEH: Get User Accounting information and Update it
    @action(detail=True, methods=['get', 'put', 'patch'],
            url_path='accounting', serializer_class=UserAccountingSerializer)
    def accounting(self, request, pk=None):
        queryset = self.get_object(pk=pk)
        if request.method in ['PUT', 'PATCH']:
            return self.custom_update(queryset, request.data, partial=(request.method == 'PATCH'))
        return self.custom_get(queryset)

    # MEH: Get User Role and Active information and Update it
    @extend_schema(tags=['Users-Role'])
    @action(detail=True, methods=['get', 'put', 'patch'],
            url_path='role', serializer_class=UserRoleSerializer, filter_backends=[None])
    def activation(self, request, pk=None):
        queryset = self.get_object(pk=pk)
        if request.method in ['PUT', 'PATCH']:
            return self.custom_update(queryset, request.data, partial=(request.method == 'PATCH'))
        return self.custom_get(queryset)

    # MEH: Get User Address list
    @extend_schema(tags=['Users-Addresses'])
    @action(detail=True, methods=['get'],
            url_path='address-list', serializer_class=AddressSerializer, filter_backends=[None])
    def address_list(self, request, pk=None):
        user = self.get_object(pk=pk)
        queryset = user.user_addresses.all()
        if not queryset.exists():
            raise NotFound(TG_DATA_EMPTY)
        return self.custom_get(queryset)

    # MEH: Add New Address (POST)
    @extend_schema(tags=['Users-Addresses'])
    @action(detail=True, methods=['post'],
            url_path='address-add', serializer_class=AddressSerializer, filter_backends=[None])
    def address_create(self, request, pk=None):
        user = self.get_object(pk=pk)
        return self.custom_create(request.data, user=user)

    # MEH: Get User Address Detail, Update and Delete it
    @extend_schema(tags=['Users-Addresses'])
    @action(detail=True, methods=['get', 'put', 'patch', 'delete'],
            url_path='address/(?P<address_id>\d+)', serializer_class=AddressSerializer, filter_backends=[None])
    def address_detail(self, request, pk=None, address_id=None):
        user = self.get_object(pk=pk)
        try:
            instance = user.user_addresses.get(pk=address_id)
        except ObjectDoesNotExist:
            raise NotFound(TG_DATA_NOT_FOUND)
        except ValueError:
            raise NotFound(TG_EXPECTED_ID_NUMBER)
        if request.method == 'DELETE':
            return self.custom_destroy(instance)
        if request.method in ['PUT', 'PATCH']:
            return self.custom_update(instance, request.data, partial=(request.method == 'PATCH'))
        return self.custom_get(instance)

    # MEH: Get User List with Filter and Back Excel Data (Download)
    @extend_schema(responses={200: UserSerializer(many=True)})
    @action(detail=False, methods=['get'],
            url_path='download')
    def download_user_list(self, request):
        # todo: reassemble Excel file later
        # Apply filters/search/order like list()
        queryset = self.filter_queryset(self.get_queryset())
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        # Header
        headers = ['ID', 'Phone Number', 'First Name', 'Last Name', 'Date Joined', 'is_active']
        ws.append(headers)
        # Set header style
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for user in queryset:
            ws.append([
                user.id,
                user.phone_number,
                user.first_name,
                user.last_name,
                user.date_joined.strftime('%Y-%m-%d %H:%M'),
                user.is_active,
            ])
        # Example: highlight inactive users
        for row in ws.iter_rows(min_row=2):
            is_active = row[5].value  # Assume you put is_active in the first column
            if not is_active:
                for cell in row:
                    cell.fill = PatternFill(start_color='FFD2D2', end_color='FFD2D2', fill_type='solid')
        # Auto width
        for col_num, column_cells in enumerate(ws.columns, 1):
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(col_num)].width = length + 2
        # Prepare HTTP response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=users.xlsx'
        wb.save(response)
        return response


# MEH: Get Introduction List and Add single object (POST) update (PUT) and protected remove (DELETE)
@extend_schema(tags=["Users-Introduction"])
class IntroductionViewSet(CustomMixinModelViewSet):
    queryset = Introduction.objects.all()
    serializer_class = IntroductionSerializer
    permission_classes = [IsAuthenticated]


# MEH: Get Role List and Add single object (POST) update (PUT) and protected remove (DELETE)
@extend_schema(tags=["Roles"])
class RoleViewSet(CustomMixinModelViewSet):
    queryset = Role.objects.all()
    serializer_class = RoleSerializer
    permission_classes = [IsAuthenticated]

    #  MEH: Delete List of object todo: work more after
    @extend_schema(
        request=BulkDeleteSerializer,
        responses={
           204: OpenApiResponse(description="Successfully deleted roles."),
           400: OpenApiResponse(description="Invalid IDs or constraint violation."),
        },
        description='Send JSON data in the body like this: {"ids": [1, 2, 3]}'
    )
    @action(detail=False, methods=['delete'], serializer_class=BulkDeleteSerializer,
            url_path='bulk-delete')
    def bulk_delete(self, request):
        return self.custom_list_destroy(request.data)
